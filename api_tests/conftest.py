import pytest
from openpyxl import load_workbook
import os
import sys
from pathlib import Path
# 将项目根目录添加到Python路径
root_dir = Path(__file__).parent.parent.absolute()
sys.path.insert(0, str(root_dir))
from tools.main_agent import main_item_chatbot


def load_test_data():
    """从Excel加载测试数据"""
    file_path = os.path.join(os.path.dirname(
        __file__), 'data', 'test.xlsx')
    wb = load_workbook(filename=file_path)
    ws = wb.active

    test_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
        if row[0] and row[1]:  # 确保A列和B列都有数据
            primary_item = row[0]
            generalizations = [g.strip()
                               for g in row[1].split('\n') if g.strip()]
            test_data.append((primary_item, generalizations))

    return test_data


@pytest.fixture(scope="module")
def chat_bot():
    """创建ChatBot实例"""
    return main_item_chatbot


@pytest.fixture(params=load_test_data())
def api_test_data(request):
    """为每个测试用例提供测试数据"""
    return request.param


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    pytest_html = item.config.pluginmanager.getplugin("html")
    outcome = yield
    report = outcome.get_result()
    extra = getattr(report, "extra", [])

    if report.when == "call":
        # 添加请求和响应信息到报告
        if "api_test_data" in item.funcargs and "chat_bot" in item.funcargs:
            primary_item, query = item.funcargs["api_test_data"]
            chat_bot = item.funcargs["chat_bot"]
            result = chat_bot.chat(query)

            extra.append(pytest_html.extras.text(
                f"测试数据:\n一级事项: {primary_item}\n查询: {query}\n"
                f"实际回答: {result['answer']}\n"
                f"响应时间: {result['duration']}秒"
            ))

        report.extra = extra
